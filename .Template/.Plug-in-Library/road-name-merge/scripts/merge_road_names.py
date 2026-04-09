#!/usr/bin/env python3
"""
도로명정보조회 엑셀 파일 통합 스크립트
Road Name Information Excel File Merge Script

Usage:
    python merge_road_names.py [province_name] [folder_path]
    python merge_road_names.py 충남 /path/to/folder
    python merge_road_names.py 전북 "C:/Users/user/Desktop/전북_도로명"

Rules applied:
  - Row 0 (총건수 행) excluded from ALL files
  - Row 1 (header)   included ONLY from the FIRST file
  - Data rows (2+)   included from ALL files
  - Output file:     [province_name]_도로명정보조회.xlsx
"""

import os
import sys
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def merge_road_name_files(folder_path: str, province_name: str) -> str:
    """
    폴더 내 모든 xlsx 파일을 통합하여 단일 엑셀 파일로 저장합니다.

    Parameters
    ----------
    folder_path   : 엑셀 파일들이 있는 폴더 경로
    province_name : 시도명 (예: 충북, 충남, 전북, 경기 …)

    Returns
    -------
    output_path : 생성된 통합 파일의 전체 경로
    """
    folder_path = os.path.abspath(folder_path)
    output_filename = f"{province_name}_도로명정보조회.xlsx"
    output_path = os.path.join(folder_path, output_filename)

    # ── 대상 파일 수집 (출력 파일 자신은 제외) ──────────────────────────────
    files = sorted([
        f for f in os.listdir(folder_path)
        if f.endswith(".xlsx") and f != output_filename
    ])

    if not files:
        raise FileNotFoundError(f"'{folder_path}' 폴더에서 xlsx 파일을 찾을 수 없습니다.")

    print(f"  발견된 파일 수: {len(files)}개")

    # ── 파일별 데이터 추출 ──────────────────────────────────────────────────
    all_data = []
    header = None

    for i, fname in enumerate(files):
        fpath = os.path.join(folder_path, fname)
        try:
            df = pd.read_excel(fpath, sheet_name=0, header=None)
        except Exception as e:
            print(f"  [경고] '{fname}' 읽기 실패 — {e}")
            continue

        # 헤더는 첫 번째 파일의 Row 1에서 한 번만 가져옴
        if i == 0:
            header = df.iloc[1].tolist()

        # Row 0(총건수), Row 1(헤더) 제외 → 순수 데이터만 추가
        data_part = df.iloc[2:].reset_index(drop=True)
        all_data.append(data_part)
        print(f"  [{i+1:02d}] {fname}  →  {len(data_part):,}행")

    if not all_data:
        raise ValueError("읽을 수 있는 데이터가 없습니다.")

    combined = pd.concat(all_data, ignore_index=True)
    total_rows = len(combined)
    print(f"\n  통합 데이터: {total_rows:,}행  ×  {len(header)}열")

    # ── openpyxl 로 서식 포함 저장 ─────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "도로명정보조회"

    # 스타일 정의
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    hdr_font  = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    hdr_fill  = PatternFill(fill_type="solid", start_color="4472C4", end_color="4472C4")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    dat_font  = Font(name="Arial", size=10)
    dat_align = Alignment(vertical="center")

    # 헤더 행 작성
    for col_idx, col_name in enumerate(header, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font   = hdr_font
        cell.fill   = hdr_fill
        cell.alignment = hdr_align
        cell.border = border

    # 데이터 행 작성
    for row_idx, row in enumerate(combined.itertuples(index=False), start=2):
        for col_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if pd.isna(val):
                cell.value = None
            elif isinstance(val, float) and val == int(val):
                cell.value = int(val)
            else:
                cell.value = val
            cell.font      = dat_font
            cell.alignment = dat_align
            cell.border    = border

    # 열 너비 (컬럼 수에 따라 자동 대응)
    default_widths = [8, 12, 20, 30, 10, 8, 14, 8, 8, 10]
    for i in range(len(header)):
        width = default_widths[i] if i < len(default_widths) else 15
        ws.column_dimensions[get_column_letter(i + 1)].width = width

    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"   # 헤더 틀 고정

    wb.save(output_path)
    print(f"\n  ✅ 저장 완료: {output_path}")
    return output_path


# ─── CLI 진입점 ────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("사용법: python merge_road_names.py [시도명] [폴더경로]")
        print("예시:   python merge_road_names.py 충남 /Users/me/충남_도로명")
        sys.exit(1)

    province = sys.argv[1]
    folder   = sys.argv[2]

    print(f"\n도로명정보조회 엑셀 통합 시작")
    print(f"  시도명  : {province}")
    print(f"  폴더경로: {folder}\n")

    try:
        result = merge_road_name_files(folder, province)
        print(f"\n완료! 생성된 파일: {result}")
    except Exception as e:
        print(f"\n[오류] {e}")
        sys.exit(1)
