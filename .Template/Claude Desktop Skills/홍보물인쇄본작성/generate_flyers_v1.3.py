#!/usr/bin/env python3
"""
매물 홍보 인쇄물 생성 스크립트 v1.3
- 매물정보.xlsx의 '매물정보' 시트에서 확인 미완료 매물만 추출
- 선택된 샘플 hwpx 형식으로 매물별 개별 hwpx 파일 생성
- hwpx 생성 완료 후 xlsx의 확인 컬럼을 '완료'로 자동 업데이트

[확인 컬럼 규칙]
  - '완료' (텍스트) : 처리 완료 → 제외 (스킵)
  - 빈 셀 (nan)     : 미처리   → 포함 (생성 대상)
  - False/0/''      : 미처리   → 포함 (Boolean 방식 호환)
  ※ 엑셀에서 '완료' 텍스트 직접 입력 방식을 사용하며, 체크박스(True/False)도 호환됨

수정이력:
  v1.3 - hwpx 생성 후 xlsx 확인 컬럼 '완료' 자동 기입 기능 추가
  v1.2 - 확인 컬럼 '완료' 텍스트 방식 명시, 한글 샘플 파일명(샘플_xxx) SAMPLE_REGISTRY 추가
  v1.1 - 비고 텍스트 색상 검정 처리, 비고 위쪽 라인 제거, 샘플 선택기 구조 추가
"""

import os
import re
import zipfile
import argparse
import pandas as pd
import openpyxl


# ─────────────────────────────────────────────
# 샘플 선택기
# 새로운 샘플 추가 시 SAMPLE_REGISTRY에만 등록하면 됩니다.
# ─────────────────────────────────────────────

SAMPLE_REGISTRY = {
    # 영문 파일명
    'sample_apartment':   '아파트',
    'sample_store':       '상가',
    'sample_land':        '토지',
    'sample_officetel':   '오피스텔',
    'sample_villa':       '빌라',
    # 한글 파일명 (샘플_xxx 형식)
    '샘플_아파트':         '아파트',
    '샘플_건물':           '상가/건물',
    '샘플_토지':           '토지',
    '샘플_오피스텔':       '오피스텔',
    '샘플_빌라':           '빌라',
}


def detect_sample_type(template_path: str) -> str:
    """템플릿 파일명에서 샘플 유형 감지. 예) 샘플_아파트.hwpx → '아파트'"""
    basename = os.path.splitext(os.path.basename(template_path))[0]
    return SAMPLE_REGISTRY.get(basename, '기타')


def list_available_templates(template_dir: str) -> list:
    """지정 디렉토리의 '샘플_' 로 시작하는 hwpx 파일 목록 반환"""
    results = []
    if not os.path.isdir(template_dir):
        return results
    for f in sorted(os.listdir(template_dir)):
        if f.startswith('샘플_') and f.endswith('.hwpx'):
            label = SAMPLE_REGISTRY.get(os.path.splitext(f)[0], '기타')
            results.append({'file': f, 'path': os.path.join(template_dir, f), 'label': label})
    return results


# ─────────────────────────────────────────────
# 금액 포맷
# ─────────────────────────────────────────────

def format_price(value) -> str:
    """억/천 단위 표시"""
    if pd.isna(value):
        return ""
    v = int(value)
    eok   = v // 10000
    cheon = (v % 10000) // 1000
    if eok > 0 and cheon > 0:
        return f"{eok}억{cheon}천"
    elif eok > 0:
        return f"{eok}억"
    elif cheon > 0:
        return f"{cheon}천"
    else:
        return str(v)


def format_price_display(row) -> str:
    """유형에 따른 가격 문자열 반환"""
    유형 = str(row['유형']).strip()
    if 유형 == '매매':
        return format_price(row.get('매매가격'))
    elif 유형 == '전세':
        return format_price(row.get('보증금'))
    elif 유형 == '월세':
        보증금 = int(row['보증금']) if pd.notna(row.get('보증금')) else 0
        월세   = int(row['월세'])   if pd.notna(row.get('월세'))   else 0
        return f"{보증금}/{월세}"
    return ""


def format_type(유형: str) -> str:
    """유형 표시 - 글자 사이 공백 삽입"""
    mapping = {'매매': '매 매', '전세': '전 세', '월세': '월 세'}
    return mapping.get(유형.strip(), 유형)


# ─────────────────────────────────────────────
# header.xml 패치
# ─────────────────────────────────────────────

def patch_header_xml(xml_bytes: bytes) -> bytes:
    """
    charPr id=9 의 textColor #FF0000 → #000000
    (비고/공인중개사 텍스트를 검정으로 변경)
    """
    xml = xml_bytes.decode('utf-8')

    def replace_color(m):
        return re.sub(r'textColor="#FF0000"', 'textColor="#000000"', m.group(0))

    xml = re.sub(r'<hh:charPr id="9"[^>]*>', replace_color, xml)
    return xml.encode('utf-8')


# ─────────────────────────────────────────────
# section0.xml 템플릿
# ─────────────────────────────────────────────
#
# borderFillIDRef 요약:
#   id=3 : 전체 실선 (상하좌우)
#   id=4 : 좌우+하단만 (topBorder=NONE) ← 비고 셀에 사용
#   id=5 : 상단+좌우만 (bottomBorder=NONE) ← 단지명 셀에 사용
#   id=6 : 좌우만 (상하 없음)
#

SECTION0_TEMPLATE = '''<hs:sec xmlns:ha="http://www.hancom.co.kr/hwpml/2011/app" xmlns:hp="http://www.hancom.co.kr/hwpml/2011/paragraph" xmlns:hp10="http://www.hancom.co.kr/hwpml/2016/paragraph" xmlns:hs="http://www.hancom.co.kr/hwpml/2011/section" xmlns:hc="http://www.hancom.co.kr/hwpml/2011/core" xmlns:hh="http://www.hancom.co.kr/hwpml/2011/head" xmlns:hhs="http://www.hancom.co.kr/hwpml/2011/history" xmlns:hm="http://www.hancom.co.kr/hwpml/2011/master-page" xmlns:hpf="http://www.hancom.co.kr/schema/2011/hpf" xmlns:dc="http://purl.org/dc/elements/1.1/" xmlns:opf="http://www.idpf.org/2007/opf/" xmlns:ooxmlchart="http://www.hancom.co.kr/hwpml/2016/ooxmlchart" xmlns:hwpunitchar="http://www.hancom.co.kr/hwpml/2016/HwpUnitChar" xmlns:epub="http://www.idpf.org/2007/ops" xmlns:config="urn:oasis:names:tc:opendocument:xmlns:config:1.0">
  <hp:p id="3121190098" paraPrIDRef="0" styleIDRef="0" pageBreak="0" columnBreak="0" merged="0">
    <hp:run charPrIDRef="0">
      <hp:secPr id="" textDirection="HORIZONTAL" spaceColumns="1134" tabStop="8000" tabStopVal="4000" tabStopUnit="HWPUNIT" outlineShapeIDRef="1" memoShapeIDRef="1" textVerticalWidthHead="0" masterPageCnt="0">
        <hp:grid lineGrid="0" charGrid="0" wonggojiFormat="0"/>
        <hp:startNum pageStartsOn="BOTH" page="0" pic="0" tbl="0" equation="0"/>
        <hp:visibility hideFirstHeader="0" hideFirstFooter="0" hideFirstMasterPage="0" border="SHOW_ALL" fill="SHOW_ALL" hideFirstPageNum="0" hideFirstEmptyLine="0" showLineNumber="0"/>
        <hp:lineNumberShape restartType="0" countBy="0" distance="0" startNumber="0"/>
        <hp:pagePr landscape="WIDELY" width="59528" height="84186" gutterType="LEFT_ONLY">
          <hp:margin header="4252" footer="4252" gutter="0" left="4251" right="4251" top="4251" bottom="2834"/>
        </hp:pagePr>
        <hp:footNotePr>
          <hp:autoNumFormat type="DIGIT" userChar="" prefixChar="" suffixChar=")" supscript="0"/>
          <hp:noteLine length="-1" type="SOLID" width="0.12 mm" color="#000000"/>
          <hp:noteSpacing betweenNotes="283" belowLine="567" aboveLine="850"/>
          <hp:numbering type="CONTINUOUS" newNum="1"/>
          <hp:placement place="EACH_COLUMN" beneathText="0"/>
        </hp:footNotePr>
        <hp:endNotePr>
          <hp:autoNumFormat type="DIGIT" userChar="" prefixChar="" suffixChar=")" supscript="0"/>
          <hp:noteLine length="14692344" type="SOLID" width="0.12 mm" color="#000000"/>
          <hp:noteSpacing betweenNotes="0" belowLine="567" aboveLine="850"/>
          <hp:numbering type="CONTINUOUS" newNum="1"/>
          <hp:placement place="END_OF_DOCUMENT" beneathText="0"/>
        </hp:endNotePr>
        <hp:pageBorderFill type="BOTH" borderFillIDRef="1" textBorder="PAPER" headerInside="0" footerInside="0" fillArea="PAPER">
          <hp:offset left="1417" right="1417" top="1417" bottom="1417"/>
        </hp:pageBorderFill>
        <hp:pageBorderFill type="EVEN" borderFillIDRef="1" textBorder="PAPER" headerInside="0" footerInside="0" fillArea="PAPER">
          <hp:offset left="1417" right="1417" top="1417" bottom="1417"/>
        </hp:pageBorderFill>
        <hp:pageBorderFill type="ODD" borderFillIDRef="1" textBorder="PAPER" headerInside="0" footerInside="0" fillArea="PAPER">
          <hp:offset left="1417" right="1417" top="1417" bottom="1417"/>
        </hp:pageBorderFill>
      </hp:secPr>
      <hp:ctrl>
        <hp:colPr id="" type="NEWSPAPER" layout="LEFT" colCount="1" sameSz="1" sameGap="0"/>
      </hp:ctrl>
    </hp:run>
    <hp:run charPrIDRef="0">
      <hp:tbl id="2131292002" zOrder="0" numberingType="TABLE" textWrap="TOP_AND_BOTTOM" textFlow="BOTH_SIDES" lock="0" dropcapstyle="None" pageBreak="CELL" repeatHeader="1" rowCnt="6" colCnt="1" cellSpacing="0" borderFillIDRef="3" noAdjust="0">
        <hp:sz width="50460" widthRelTo="ABSOLUTE" height="67662" heightRelTo="ABSOLUTE" protect="0"/>
        <hp:pos treatAsChar="1" affectLSpacing="0" flowWithText="1" allowOverlap="0" holdAnchorAndSO="0" vertRelTo="PARA" horzRelTo="COLUMN" vertAlign="TOP" horzAlign="LEFT" vertOffset="0" horzOffset="0"/>
        <hp:outMargin left="283" right="283" top="283" bottom="283"/>
        <hp:inMargin left="510" right="510" top="141" bottom="141"/>
        <hp:tr>
          <hp:tc name="" header="0" hasMargin="0" protect="0" editable="0" dirty="0" borderFillIDRef="5">
            <hp:subList id="" textDirection="HORIZONTAL" lineWrap="BREAK" vertAlign="CENTER" linkListIDRef="0" linkListNextIDRef="0" textWidth="0" textHeight="0" hasTextRef="0" hasNumRef="0">
              <hp:p id="0" paraPrIDRef="20" styleIDRef="0" pageBreak="0" columnBreak="0" merged="0">
                <hp:run charPrIDRef="12"><hp:t>{단지명}</hp:t></hp:run>
                <hp:run charPrIDRef="10"/>
                <hp:linesegarray><hp:lineseg textpos="0" vertpos="0" vertsize="6900" textheight="6900" baseline="5865" spacing="4140" horzpos="0" horzsize="49440" flags="393216"/></hp:linesegarray>
              </hp:p>
            </hp:subList>
            <hp:cellAddr colAddr="0" rowAddr="0"/><hp:cellSpan colSpan="1" rowSpan="1"/>
            <hp:cellSz width="50460" height="12255"/>
            <hp:cellMargin left="510" right="510" top="141" bottom="141"/>
          </hp:tc>
        </hp:tr>
        <hp:tr>
          <hp:tc name="" header="0" hasMargin="0" protect="0" editable="0" dirty="0" borderFillIDRef="6">
            <hp:subList id="" textDirection="HORIZONTAL" lineWrap="BREAK" vertAlign="CENTER" linkListIDRef="0" linkListNextIDRef="0" textWidth="0" textHeight="0" hasTextRef="0" hasNumRef="0">
              <hp:p id="0" paraPrIDRef="20" styleIDRef="0" pageBreak="0" columnBreak="0" merged="0">
                <hp:run charPrIDRef="8"><hp:t>{면적}</hp:t></hp:run>
                <hp:linesegarray><hp:lineseg textpos="0" vertpos="0" vertsize="6000" textheight="6000" baseline="5100" spacing="3600" horzpos="0" horzsize="49440" flags="393216"/></hp:linesegarray>
              </hp:p>
            </hp:subList>
            <hp:cellAddr colAddr="0" rowAddr="1"/><hp:cellSpan colSpan="1" rowSpan="1"/>
            <hp:cellSz width="50460" height="12255"/>
            <hp:cellMargin left="510" right="510" top="141" bottom="141"/>
          </hp:tc>
        </hp:tr>
        <hp:tr>
          <hp:tc name="" header="0" hasMargin="0" protect="0" editable="0" dirty="0" borderFillIDRef="6">
            <hp:subList id="" textDirection="HORIZONTAL" lineWrap="BREAK" vertAlign="CENTER" linkListIDRef="0" linkListNextIDRef="0" textWidth="0" textHeight="0" hasTextRef="0" hasNumRef="0">
              <hp:p id="0" paraPrIDRef="20" styleIDRef="0" pageBreak="0" columnBreak="0" merged="0">
                <hp:run charPrIDRef="11"><hp:t>{유형}</hp:t></hp:run>
                <hp:linesegarray><hp:lineseg textpos="0" vertpos="0" vertsize="7000" textheight="7000" baseline="5950" spacing="4200" horzpos="0" horzsize="49440" flags="393216"/></hp:linesegarray>
              </hp:p>
            </hp:subList>
            <hp:cellAddr colAddr="0" rowAddr="2"/><hp:cellSpan colSpan="1" rowSpan="1"/>
            <hp:cellSz width="50460" height="12255"/>
            <hp:cellMargin left="510" right="510" top="141" bottom="141"/>
          </hp:tc>
        </hp:tr>
        <hp:tr>
          <hp:tc name="" header="0" hasMargin="0" protect="0" editable="0" dirty="0" borderFillIDRef="6">
            <hp:subList id="" textDirection="HORIZONTAL" lineWrap="BREAK" vertAlign="CENTER" linkListIDRef="0" linkListNextIDRef="0" textWidth="0" textHeight="0" hasTextRef="0" hasNumRef="0">
              <hp:p id="0" paraPrIDRef="20" styleIDRef="0" pageBreak="0" columnBreak="0" merged="0">
                <hp:run charPrIDRef="8"><hp:t>{가격}</hp:t></hp:run>
                <hp:linesegarray><hp:lineseg textpos="0" vertpos="0" vertsize="6000" textheight="6000" baseline="5100" spacing="3600" horzpos="0" horzsize="49440" flags="393216"/></hp:linesegarray>
              </hp:p>
            </hp:subList>
            <hp:cellAddr colAddr="0" rowAddr="3"/><hp:cellSpan colSpan="1" rowSpan="1"/>
            <hp:cellSz width="50460" height="12255"/>
            <hp:cellMargin left="510" right="510" top="141" bottom="141"/>
          </hp:tc>
        </hp:tr>
        <hp:tr>
          <!-- 비고: borderFillIDRef=4 (topBorder=NONE → 위쪽 라인 없음) -->
          <hp:tc name="" header="0" hasMargin="0" protect="0" editable="0" dirty="0" borderFillIDRef="4">
            <hp:subList id="" textDirection="HORIZONTAL" lineWrap="BREAK" vertAlign="CENTER" linkListIDRef="0" linkListNextIDRef="0" textWidth="0" textHeight="0" hasTextRef="0" hasNumRef="0">
              <hp:p id="0" paraPrIDRef="20" styleIDRef="0" pageBreak="0" columnBreak="0" merged="0">
                <hp:run charPrIDRef="9"><hp:t>{비고}</hp:t></hp:run>
                <hp:linesegarray><hp:lineseg textpos="0" vertpos="0" vertsize="3000" textheight="3000" baseline="2550" spacing="1800" horzpos="0" horzsize="49440" flags="393216"/></hp:linesegarray>
              </hp:p>
            </hp:subList>
            <hp:cellAddr colAddr="0" rowAddr="4"/><hp:cellSpan colSpan="1" rowSpan="1"/>
            <hp:cellSz width="50460" height="12255"/>
            <hp:cellMargin left="510" right="510" top="141" bottom="141"/>
          </hp:tc>
        </hp:tr>
        <hp:tr>
          <hp:tc name="" header="0" hasMargin="0" protect="0" editable="0" dirty="0" borderFillIDRef="3">
            <hp:subList id="" textDirection="HORIZONTAL" lineWrap="BREAK" vertAlign="CENTER" linkListIDRef="0" linkListNextIDRef="0" textWidth="0" textHeight="0" hasTextRef="0" hasNumRef="0">
              <hp:p id="0" paraPrIDRef="20" styleIDRef="0" pageBreak="0" columnBreak="0" merged="0">
                <hp:run charPrIDRef="9"><hp:t>{공인중개사}</hp:t></hp:run>
                <hp:linesegarray><hp:lineseg textpos="0" vertpos="0" vertsize="3000" textheight="3000" baseline="2550" spacing="1800" horzpos="0" horzsize="49440" flags="393216"/></hp:linesegarray>
              </hp:p>
            </hp:subList>
            <hp:cellAddr colAddr="0" rowAddr="5"/><hp:cellSpan colSpan="1" rowSpan="1"/>
            <hp:cellSz width="50460" height="6387"/>
            <hp:cellMargin left="510" right="510" top="141" bottom="141"/>
          </hp:tc>
        </hp:tr>
      </hp:tbl>
      <hp:t/>
    </hp:run>
    <hp:linesegarray>
      <hp:lineseg textpos="0" vertpos="0" vertsize="68228" textheight="68228" baseline="57994" spacing="600" horzpos="0" horzsize="51024" flags="393216"/>
    </hp:linesegarray>
  </hp:p>
</hs:sec>'''


def make_prvtext(단지명, 면적, 유형, 가격, 비고, 공인중개사):
    return f"<{단지명}>\r\n<{면적}>\r\n<{유형}>\r\n<{가격}>\r\n<{비고}>\r\n<{공인중개사}>\r\n"


# ─────────────────────────────────────────────
# hwpx 생성
# ─────────────────────────────────────────────

EMPTY_PNG = bytes([
    0x89,0x50,0x4E,0x47,0x0D,0x0A,0x1A,0x0A,
    0x00,0x00,0x00,0x0D,0x49,0x48,0x44,0x52,
    0x00,0x00,0x00,0x01,0x00,0x00,0x00,0x01,
    0x08,0x06,0x00,0x00,0x00,0x1F,0x15,0xC4,
    0x89,0x00,0x00,0x00,0x0A,0x49,0x44,0x41,
    0x54,0x78,0x9C,0x62,0x00,0x01,0x00,0x00,
    0x05,0x00,0x01,0x0D,0x0A,0x2D,0xB4,0x00,
    0x00,0x00,0x00,0x49,0x45,0x4E,0x44,0xAE,
    0x42,0x60,0x82
])


def create_hwpx(template_path: str, output_path: str, data: dict):
    section_xml = SECTION0_TEMPLATE.format(
        단지명=_esc(data['단지명']), 면적=_esc(data['면적']),
        유형=_esc(data['유형']),   가격=_esc(data['가격']),
        비고=_esc(data['비고']),   공인중개사=_esc(data['공인중개사']),
    )
    prvtext = make_prvtext(
        data['단지명'], data['면적'], data['유형'],
        data['가격'], data['비고'], data['공인중개사']
    )

    with zipfile.ZipFile(template_path, 'r') as tmpl:
        file_contents = {}
        for name in tmpl.namelist():
            raw = tmpl.read(name)
            if name == 'Contents/header.xml':
                raw = patch_header_xml(raw)
            if name not in ('Contents/section0.xml', 'Preview/PrvText.txt', 'Preview/PrvImage.png'):
                file_contents[name] = raw

    with zipfile.ZipFile(output_path, 'w', zipfile.ZIP_DEFLATED) as zout:
        if 'mimetype' in file_contents:
            zout.writestr(zipfile.ZipInfo('mimetype'), file_contents['mimetype'])
        for name, content in file_contents.items():
            if name == 'mimetype':
                continue
            zout.writestr(name, content)
        zout.writestr('Contents/section0.xml', section_xml.encode('utf-8'))
        zout.writestr('Preview/PrvText.txt',   prvtext.encode('utf-8'))
        zout.writestr('Preview/PrvImage.png',  EMPTY_PNG)


def _esc(s: str) -> str:
    return str(s).replace('&','&amp;').replace('<','&lt;').replace('>','&gt;').replace('"','&quot;')


# ─────────────────────────────────────────────
# 메인
# ─────────────────────────────────────────────

def generate_flyers(xlsx_path: str, template_path: str, outdir: str,
                    공인중개사: str = '크로바공인중개사사무소'):
    os.makedirs(outdir, exist_ok=True)

    sample_type = detect_sample_type(template_path)
    print(f"  템플릿 유형: {sample_type}  ({os.path.basename(template_path)})")

    df = pd.read_excel(xlsx_path, sheet_name='매물정보')
    # 확인 컬럼 필터링:
    #   '완료' 텍스트가 있는 행 → 제외 (이미 처리된 매물)
    #   빈 셀(nan), False, 0, '' → 포함 (미처리 매물, 생성 대상)
    미확인 = df[df['확인'].astype(str).str.lower().isin(['false', '0', 'nan', ''])]

    # openpyxl로 xlsx 열기 (확인 컬럼 업데이트용)
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb['매물정보']

    # 헤더 행에서 '확인' 컬럼 위치 찾기
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=False))
    확인_col_idx = None
    for cell in header_row:
        if cell.value == '확인':
            확인_col_idx = cell.column
            break

    generated = []
    # 미확인 행의 원본 인덱스(0-based) → xlsx 행 번호(2-based, 헤더 제외) 매핑
    for df_idx, row in 미확인.iterrows():
        단지명   = str(row['단지명']).strip()
        면적     = str(row['면적']).strip()
        유형_raw = str(row['유형']).strip()
        유형     = format_type(유형_raw)
        가격     = format_price_display(row)
        비고     = str(row['비고']).strip() if pd.notna(row.get('비고')) else ''
        동       = str(row.get('동', '')).strip()
        호       = str(row.get('호', '')).strip()

        safe_name   = f"{단지명}_{동}_{호}_{유형_raw}.hwpx".replace(' ', '_').replace('/', '_')
        output_path = os.path.join(outdir, safe_name)

        create_hwpx(template_path, output_path, {
            '단지명': 단지명, '면적': 면적, '유형': 유형,
            '가격': 가격, '비고': 비고, '공인중개사': 공인중개사,
        })
        generated.append(output_path)
        print(f"  생성: {safe_name}  [{단지명} / {면적} / {유형} / {가격}]")

        # xlsx 확인 컬럼 → '완료' 기입 (df_idx는 0-based, xlsx는 헤더=1행이므로 +2)
        if 확인_col_idx is not None:
            ws.cell(row=df_idx + 2, column=확인_col_idx).value = '완료'

    # 변경된 xlsx 저장
    wb.save(xlsx_path)
    print(f"  ✅ 매물정보.xlsx 확인 컬럼 '완료' 업데이트 완료")

    print(f"\n총 {len(generated)}개 파일 생성 완료 → {outdir}")
    return generated


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='매물 홍보 인쇄물 hwpx 생성 v1.1')
    parser.add_argument('--xlsx',     required=True, help='매물정보.xlsx 경로')
    parser.add_argument('--template', required=True, help='샘플 hwpx 경로 (예: 샘플_아파트.hwpx)')
    parser.add_argument('--outdir',   required=True, help='출력 디렉토리')
    parser.add_argument('--agency',   default='크로바공인중개사사무소', help='공인중개사 사무소명')
    parser.add_argument('--list-templates', metavar='DIR',
                        help='지정 디렉토리의 사용 가능한 샘플 목록 출력 후 종료')
    args = parser.parse_args()

    if args.list_templates:
        templates = list_available_templates(args.list_templates)
        if templates:
            print("사용 가능한 샘플 템플릿:")
            for t in templates:
                print(f"  {t['file']}  →  {t['label']} 유형")
        else:
            print(f"'{args.list_templates}' 에서 샘플 파일을 찾을 수 없습니다.")
    else:
        generate_flyers(args.xlsx, args.template, args.outdir, args.agency)
